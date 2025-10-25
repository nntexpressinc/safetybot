#!/usr/bin/env python3
"""
Safety Bot - GoMotive APIs to Telegram Bot with Excel Storage
Monitors speeding and driver performance events with daily Excel exports
"""

import os
import sys
import time
import json
import asyncio
import logging
import requests
import tempfile
import uuid
import signal
import atexit
import threading
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Set, Tuple
from dotenv import load_dotenv
import schedule
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from telegram import Bot, InputMediaVideo
from telegram.error import TelegramError, NetworkError, TimedOut
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load environment variables
load_dotenv()

# Configure logging with proper encoding
def setup_logging():
    """Set up logging with Unicode support and fallback"""
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    file_handler = logging.FileHandler('safetybot.log', encoding='utf-8')
    file_handler.setFormatter(formatter)
    
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass
    
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

logger = setup_logging()

def log_safe(level, message):
    """Safe logging function that handles Unicode issues"""
    try:
        getattr(logger, level)(message)
    except UnicodeEncodeError:
        safe_message = message.encode('ascii', 'ignore').decode('ascii')
        getattr(logger, level)(safe_message)

class SafetyBot:
    """Main bot class for monitoring GoMotive APIs and storing/sending Excel reports"""
    
    ALLOWED_EVENT_TYPES = [
        'hard_brake', 'crash', 'seat_belt_violation', 'stop_sign_violation',
        'distraction', 'unsafe_lane_change'
    ]
    
    # Constraints
    MAX_VIDEO_SIZE_MB = 20
    VIDEO_DOWNLOAD_TIMEOUT = 90
    TELEGRAM_SEND_TIMEOUT = 30
    MAX_QUEUE_SIZE = 50
    
    def __init__(self):
        """Initialize the bot with configuration from environment variables"""
        self.api_key = os.getenv('API_KEY')
        self.api_base_url = os.getenv('API_BASE_URL')
        self.telegram_token = os.getenv('TELEGRAM_BOT_TOKEN')
        self.chat_id = os.getenv('TELEGRAM_CHAT_ID')
        self.check_interval = int(os.getenv('CHECK_INTERVAL', 300))
        
        # Health monitoring
        self.last_successful_check = None
        self.consecutive_failures = 0
        self.max_consecutive_failures = 5
        self.critical_alert_sent = False
        
        # Event tracking files
        self.last_performance_event_file = 'last_performance_event_id.txt'
        self.last_speeding_event_file = 'last_speeding_event_id.txt'
        
        self.performance_event_files = {
            event_type: f'last_{event_type}_event_id.txt'
            for event_type in self.ALLOWED_EVENT_TYPES
        }
        
        # Data storage - today's events
        self.today_events = []
        self.data_lock = threading.Lock()
        
        # Deduplication within current cycle
        self.processed_event_ids: Set[int] = set()
        
        # Concurrency and shutdown control
        self.is_processing = False
        self.running = True
        self.shutdown_lock = threading.Lock()
        
        # Validate and initialize
        self._validate_config()
        self.telegram_bot = Bot(token=self.telegram_token)
        self._init_session()
        
        # Register cleanup handlers
        signal.signal(signal.SIGTERM, self._signal_handler)
        signal.signal(signal.SIGINT, self._signal_handler)
        atexit.register(self._cleanup)
        
        logger.info("SafetyBot initialized successfully")
    
    def _signal_handler(self, signum, frame):
        """Handle shutdown signals gracefully"""
        log_safe('info', f"Received signal {signum}, initiating graceful shutdown...")
        with self.shutdown_lock:
            self.running = False
    
    def _cleanup(self):
        """Clean up resources on exit"""
        try:
            with self.shutdown_lock:
                self.running = False
            
            if hasattr(self, 'session'):
                self.session.close()
            logger.info("Cleanup completed successfully")
        except Exception as e:
            logger.error(f"Error during cleanup: {e}")
    
    def _init_session(self):
        """Initialize requests session with retry strategy"""
        self.headers = {
            'accept': 'application/json',
            'x-api-key': self.api_key,
            'User-Agent': 'SafetyBot/2.1'
        }
        
        self.session = requests.Session()
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "OPTIONS"]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy, pool_connections=10, pool_maxsize=10)
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)
    
    def _validate_config(self):
        """Validate required configuration"""
        required_vars = ['API_KEY', 'API_BASE_URL', 'TELEGRAM_BOT_TOKEN', 'TELEGRAM_CHAT_ID']
        missing_vars = [var for var in required_vars if not os.getenv(var)]
        
        if missing_vars:
            raise ValueError(f"Missing required environment variables: {', '.join(missing_vars)}")
    
    def _has_allowed_severity(self, event: Dict[str, Any]) -> bool:
        """Check if event has allowed severity"""
        metadata = event.get('metadata', {})
        severity = metadata.get('severity')
        return severity in ['medium', 'high', 'critical']
    
    def _is_already_processed(self, event_id: int) -> bool:
        """Check if event was already processed in this cycle"""
        return event_id in self.processed_event_ids
    
    def _mark_processed(self, event_id: int):
        """Mark event as processed in this cycle"""
        self.processed_event_ids.add(event_id)
    
    def _reset_cycle_tracking(self):
        """Reset deduplication tracking for new cycle"""
        self.processed_event_ids.clear()
    
    def get_last_processed_event_id(self, event_type='performance') -> int:
        """Get the ID of the last processed event"""
        try:
            file_path = self.last_performance_event_file if event_type == 'performance' else self.last_speeding_event_file
            if os.path.exists(file_path):
                with open(file_path, 'r') as f:
                    return int(f.read().strip())
        except (FileNotFoundError, ValueError) as e:
            logger.warning(f"Could not read last {event_type} event ID: {e}")
        return 0
    
    def save_last_processed_event_id(self, event_id: int, event_type='performance'):
        """Save the ID of the last processed event"""
        try:
            file_path = self.last_performance_event_file if event_type == 'performance' else self.last_speeding_event_file
            with open(file_path, 'w') as f:
                f.write(str(event_id))
            logger.info(f"[ID_SAVED] {event_type}: {event_id}")
        except Exception as e:
            logger.error(f"Could not save last {event_type} event ID: {e}")
    
    def get_last_processed_event_id_for_type(self, event_type: str) -> int:
        """Get the last processed event ID for a specific performance event type"""
        file_path = self.performance_event_files.get(event_type, self.last_performance_event_file)
        try:
            if os.path.exists(file_path):
                with open(file_path, 'r') as f:
                    return int(f.read().strip())
        except (FileNotFoundError, ValueError) as e:
            logger.warning(f"Could not read last {event_type} event ID: {e}")
        return 0
    
    def save_last_processed_event_id_for_type(self, event_id: int, event_type: str):
        """Save the last processed event ID for a specific performance event type"""
        file_path = self.performance_event_files.get(event_type, self.last_performance_event_file)
        try:
            with open(file_path, 'w') as f:
                f.write(str(event_id))
            logger.info(f"[ID_SAVED] {event_type}: {event_id}")
        except IOError as e:
            logger.error(f"Error saving last {event_type} event ID: {e}")
    
    def fetch_speeding_events(self) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
        """Fetch speeding events from GoMotive v1 API. Returns (events, error_code)"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                params = {'per_page': '25', 'page_no': '1'}
                url = f"{self.api_base_url.replace('v2', 'v1')}/speeding_events"
                
                logger.info(f"Fetching speeding events (attempt {attempt + 1}/{max_retries})")
                response = self.session.get(url, headers=self.headers, params=params, timeout=45)
                
                if response.status_code == 401:
                    logger.error("API Authentication failed (401) - Invalid API key")
                    return None, "AUTH_ERROR"
                elif response.status_code == 403:
                    logger.error("API Permission denied (403)")
                    return None, "PERMISSION_ERROR"
                elif response.status_code == 404:
                    logger.error("API Endpoint not found (404)")
                    return None, "NOT_FOUND"
                
                response.raise_for_status()
                
                if not response.content:
                    logger.warning("Empty response from speeding events API")
                    return [], None
                
                data = response.json()
                events = data.get('speeding_events', [])
                
                if not isinstance(events, list):
                    logger.warning("Unexpected data structure in speeding events response")
                    return [], None
                
                logger.info(f"Successfully fetched {len(events)} speeding events")
                return events, None
                
            except requests.exceptions.Timeout:
                logger.warning(f"Timeout fetching speeding events (attempt {attempt + 1}/{max_retries})")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
            except requests.exceptions.RequestException as e:
                logger.error(f"Request error fetching speeding events (attempt {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
            except json.JSONDecodeError as e:
                logger.error(f"JSON decode error in speeding events response: {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
            except Exception as e:
                logger.error(f"Unexpected error fetching speeding events: {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
        
        logger.error(f"Failed to fetch speeding events after {max_retries} attempts")
        return None, "FETCH_ERROR"
    
    def fetch_driver_performance_events(self) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
        """Fetch driver performance events from GoMotive v2 API. Returns (events, error_code)"""
        all_events = []
        first_error = None
        
        for event_type in self.ALLOWED_EVENT_TYPES:
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    params = {
                        'event_types': event_type,
                        'media_required': 'true',
                        'per_page': '25',
                        'page_no': '1'
                    }
                    
                    url = f"{self.api_base_url}/driver_performance_events"
                    logger.info(f"Fetching {event_type} events (attempt {attempt + 1}/{max_retries})")
                    response = self.session.get(url, headers=self.headers, params=params, timeout=45)
                    
                    if response.status_code == 401:
                        logger.error(f"API Authentication failed (401) for {event_type}")
                        if not first_error:
                            first_error = "AUTH_ERROR"
                        break
                    elif response.status_code == 403:
                        logger.error(f"API Permission denied (403) for {event_type}")
                        if not first_error:
                            first_error = "PERMISSION_ERROR"
                        break
                    
                    response.raise_for_status()
                    
                    if not response.content:
                        logger.warning(f"Empty response from {event_type} events API")
                        break
                    
                    data = response.json()
                    events = data.get('driver_performance_events', [])
                    
                    if not isinstance(events, list):
                        logger.warning(f"Unexpected data structure in {event_type} events response")
                        break
                    
                    logger.info(f"Found {len(events)} {event_type} events")
                    all_events.extend(events)
                    break
                    
                except requests.exceptions.Timeout:
                    logger.warning(f"Timeout fetching {event_type} events (attempt {attempt + 1}/{max_retries})")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)
                except requests.exceptions.RequestException as e:
                    logger.error(f"Request error fetching {event_type} events: {e}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)
                except Exception as e:
                    logger.error(f"Unexpected error fetching {event_type} events: {e}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)
            
            time.sleep(1)
        
        logger.info(f"Successfully fetched {len(all_events)} total performance events")
        return all_events, first_error
    
    def filter_new_speeding_events(self, events: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Filter new speeding events with allowed severity. Saves ID immediately."""
        last_event_id = self.get_last_processed_event_id('speeding')
        new_events = []
        
        for event_data in events:
            event = event_data.get('speeding_event', {})
            event_id = event.get('id', 0)
            
            if event_id > last_event_id and self._has_allowed_severity(event):
                if not self._is_already_processed(event_id):
                    new_events.append(event)
        
        new_events.sort(key=lambda x: x.get('id', 0))
        logger.info(f"Found {len(new_events)} new speeding events (last ID: {last_event_id})")
        
        if new_events:
            max_id = max([e.get('id', 0) for e in new_events])
            self.save_last_processed_event_id(max_id, 'speeding')
        
        return new_events
    
    def filter_new_performance_events(self, events: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Filter new performance events per event type. Saves IDs immediately."""
        new_events = []
        events_by_type = {}
        
        for event_data in events:
            event = event_data.get('driver_performance_event', {})
            event_type = event.get('type', '')
            if event_type in self.ALLOWED_EVENT_TYPES:
                if event_type not in events_by_type:
                    events_by_type[event_type] = []
                events_by_type[event_type].append(event)
        
        for event_type, type_events in events_by_type.items():
            last_event_id = self.get_last_processed_event_id_for_type(event_type)
            type_new_events = []
            
            for event in type_events:
                event_id = event.get('id', 0)
                if event_id > last_event_id and self._has_allowed_severity(event):
                    if not self._is_already_processed(event_id):
                        type_new_events.append(event)
            
            type_new_events.sort(key=lambda x: x.get('id', 0))
            logger.info(f"Found {len(type_new_events)} new {event_type} events (last ID: {last_event_id})")
            
            if type_new_events:
                max_id_for_type = max([e.get('id', 0) for e in type_new_events])
                self.save_last_processed_event_id_for_type(max_id_for_type, event_type)
            
            new_events.extend(type_new_events)
        
        new_events.sort(key=lambda x: x.get('id', 0))
        return new_events
    
    def format_time(self, time_str: str, latitude: Optional[float] = None, longitude: Optional[float] = None) -> str:
        """Format time string to readable format with timezone conversion"""
        try:
            import pytz
            
            dt_utc = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
            dt_utc = dt_utc.astimezone(pytz.UTC)
            
            tz_local = None
            if longitude is not None:
                if longitude >= -125 and longitude < -120:
                    tz_local = pytz.timezone('America/Los_Angeles')
                elif longitude >= -120 and longitude < -104:
                    tz_local = pytz.timezone('America/Denver')
                elif longitude >= -104 and longitude < -87:
                    tz_local = pytz.timezone('America/Chicago')
                elif longitude >= -87 and longitude <= -60:
                    tz_local = pytz.timezone('America/New_York')
                else:
                    tz_local = pytz.timezone('America/Los_Angeles')
            else:
                tz_local = pytz.timezone('America/Los_Angeles')
            
            dt_local = dt_utc.astimezone(tz_local)
            dt_local = dt_local - timedelta(hours=1)
            
            formatted = dt_local.strftime('%m/%d/%Y %I:%M %p')
            return formatted
            
        except Exception as e:
            logger.error(f"[TZ_ERROR] Error formatting time '{time_str}': {e}")
            return time_str
    
    def store_speeding_event(self, event: Dict[str, Any]):
        """Store speeding event in memory"""
        try:
            driver = event.get('driver', {})
            driver_name = f"{driver.get('first_name', '')} {driver.get('last_name', '')}".strip() or 'Unknown'
            
            start_lat = event.get('start_lat')
            start_lon = event.get('start_lon')
            date_time = self.format_time(event.get('start_time', ''), start_lat, start_lon)
            
            min_speed = event.get('min_vehicle_speed', 0)
            max_speed = event.get('max_vehicle_speed', 0)
            avg_exceeded = event.get('avg_over_speed_in_kph', 0)
            
            min_speed_mph = round(min_speed * 0.621371, 1) if min_speed else 0
            max_speed_mph = round(max_speed * 0.621371, 1) if max_speed else 0
            avg_exceeded_mph = round(avg_exceeded * 0.621371, 1) if avg_exceeded else 0
            
            speed_range = f"{min_speed_mph}–{max_speed_mph} mph"
            severity = event.get('metadata', {}).get('severity', 'unknown')
            
            event_record = {
                'event_type': 'Speeding',
                'driver_name': driver_name,
                'date_time': date_time,
                'speed_range': speed_range,
                'exceeded_by': f"+{avg_exceeded_mph} mph",
                'severity': severity
            }
            
            with self.data_lock:
                self.today_events.append(event_record)
            
            logger.info(f"[STORED] Speeding event: {driver_name}")
            
        except Exception as e:
            logger.error(f"[STORE_ERROR] Error storing speeding event: {e}")
    
    def store_performance_event(self, event: Dict[str, Any]):
        """Store performance event in memory"""
        try:
            event_type = event.get('type', 'Unknown').replace('_', ' ').title()
            driver = event.get('driver', {})
            driver_name = f"{driver.get('first_name', '')} {driver.get('last_name', '')}".strip() or 'Unknown'
            
            end_lat = event.get('end_lat')
            end_lon = event.get('end_lon')
            end_time = self.format_time(event.get('end_time', ''), end_lat, end_lon)
            
            severity = event.get('metadata', {}).get('severity', 'unknown')
            
            event_record = {
                'event_type': event_type,
                'driver_name': driver_name,
                'date_time': end_time,
                'speed_range': '',
                'exceeded_by': '',
                'severity': severity
            }
            
            with self.data_lock:
                self.today_events.append(event_record)
            
            logger.info(f"[STORED] Performance event: {event_type} - {driver_name}")
            
        except Exception as e:
            logger.error(f"[STORE_ERROR] Error storing performance event: {e}")
    
    def create_excel_file(self, events: Optional[List[Dict[str, str]]] = None) -> Optional[str]:
        """Create Excel file with events data. Returns file path or None."""
        try:
            if events is None:
                with self.data_lock:
                    events = list(self.today_events)
            
            if not events:
                logger.warning("[EXCEL] No events to export")
                return None
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Safety Events"
            
            # Define headers
            headers = [
                'Event Type',
                'Driver Name',
                'Date & Time',
                'Speed Range',
                'Exceeded By',
                'Severity'
            ]
            
            # Add headers with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Add data rows
            for row_idx, event in enumerate(events, 2):
                ws.cell(row=row_idx, column=1).value = event.get('event_type', '')
                ws.cell(row=row_idx, column=2).value = event.get('driver_name', '')
                ws.cell(row=row_idx, column=3).value = event.get('date_time', '')
                ws.cell(row=row_idx, column=4).value = event.get('speed_range', '')
                ws.cell(row=row_idx, column=5).value = event.get('exceeded_by', '')
                ws.cell(row=row_idx, column=6).value = event.get('severity', '')
                
                # Apply borders to data rows
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).border = border
                    ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="left", vertical="center")
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            
            # Save file
            temp_filename = f"Safety_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            temp_file_path = os.path.join(tempfile.gettempdir(), temp_filename)
            wb.save(temp_file_path)
            
            logger.info(f"[EXCEL] Created Excel file: {temp_file_path} with {len(events)} events")
            return temp_file_path
            
        except Exception as e:
            logger.error(f"[EXCEL_ERROR] Error creating Excel file: {e}")
            return None
    
    async def send_excel_file(self, file_path: str, caption: str = "Daily Safety Report"):
        """Send Excel file to Telegram"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                with open(file_path, 'rb') as f:
                    await self.telegram_bot.send_document(
                        chat_id=self.chat_id,
                        document=f,
                        caption=caption,
                        read_timeout=self.TELEGRAM_SEND_TIMEOUT,
                        write_timeout=self.TELEGRAM_SEND_TIMEOUT
                    )
                logger.info(f"[SENT] Excel file sent: {caption}")
                return True
                
            except (NetworkError, TimedOut) as e:
                logger.warning(f"[RETRY] Network error (attempt {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(2 ** attempt)
            except TelegramError as e:
                logger.error(f"[ERROR] Telegram error: {e}")
                return False
            except Exception as e:
                logger.error(f"[ERROR] Unexpected error (attempt {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(2 ** attempt)
        
        logger.error(f"[FAILED] Could not send Excel file")
        return False
    
    async def send_speeding_event_to_telegram(self, event: Dict[str, Any]) -> bool:
        """Send speeding event to Telegram"""
        try:
            driver = event.get('driver', {})
            driver_name = f"{driver.get('first_name', '')} {driver.get('last_name', '')}".strip() or 'Unknown'
            
            vehicle_number = event.get('vehicle', {}).get('number', 'N/A')
            start_lat = event.get('start_lat')
            start_lon = event.get('start_lon')
            date_time = self.format_time(event.get('start_time', ''), start_lat, start_lon)
            
            min_speed = event.get('min_vehicle_speed', 0)
            max_speed = event.get('max_vehicle_speed', 0)
            avg_exceeded = event.get('avg_over_speed_in_kph', 0)
            
            min_speed_mph = round(min_speed * 0.621371, 1) if min_speed else 0
            max_speed_mph = round(max_speed * 0.621371, 1) if max_speed else 0
            avg_exceeded_mph = round(avg_exceeded * 0.621371, 1) if avg_exceeded else 0
            
            severity = event.get('metadata', {}).get('severity', 'unknown')
            
            message = f"""Speeding Alert
Driver: {driver_name}
Vehicle: {vehicle_number}
{date_time}
Speed Range: {min_speed_mph}–{max_speed_mph} mph
Exceeded By: +{avg_exceeded_mph} mph
Severity: {severity}"""
            
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    await self.telegram_bot.send_message(
                        chat_id=self.chat_id,
                        text=message,
                        parse_mode='Markdown',
                        read_timeout=self.TELEGRAM_SEND_TIMEOUT,
                        write_timeout=self.TELEGRAM_SEND_TIMEOUT
                    )
                    logger.info(f"[SENT] Speeding event {event.get('id')}")
                    return True
                    
                except (NetworkError, TimedOut) as e:
                    logger.warning(f"[RETRY] Network error (attempt {attempt + 1}/{max_retries}): {e}")
                    if attempt < max_retries - 1:
                        await asyncio.sleep(2 ** attempt)
                except TelegramError as e:
                    logger.error(f"[ERROR] Telegram API error: {e}")
                    return False
                except Exception as e:
                    logger.error(f"[ERROR] Unexpected error (attempt {attempt + 1}/{max_retries}): {e}")
                    if attempt < max_retries - 1:
                        await asyncio.sleep(2 ** attempt)
        
        except Exception as e:
            logger.error(f"[FORMAT_ERROR] Error formatting speeding message: {e}")
        
        return False
    
    async def send_performance_event_to_telegram(self, event: Dict[str, Any]) -> bool:
        """Send performance event to Telegram"""
        try:
            event_type = event.get('type', 'Unknown').replace('_', ' ').title()
            vehicle_number = event.get('vehicle', {}).get('number', 'N/A')
            driver = event.get('driver', {})
            driver_name = f"{driver.get('first_name', '')} {driver.get('last_name', '')}".strip() or 'Unknown'
            
            end_lat = event.get('end_lat')
            end_lon = event.get('end_lon')
            end_time = self.format_time(event.get('end_time', ''), end_lat, end_lon)
            
            severity = event.get('metadata', {}).get('severity', 'unknown')
            
            message = f"""{event_type}
Driver: {driver_name}
Vehicle: {vehicle_number}
{end_time}
Severity: {severity}"""
            
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    await self.telegram_bot.send_message(
                        chat_id=self.chat_id,
                        text=message,
                        parse_mode='Markdown',
                        read_timeout=self.TELEGRAM_SEND_TIMEOUT,
                        write_timeout=self.TELEGRAM_SEND_TIMEOUT
                    )
                    logger.info(f"[SENT] Performance event {event.get('id')}")
                    return True
                    
                except (NetworkError, TimedOut) as e:
                    logger.warning(f"[RETRY] Network error (attempt {attempt + 1}/{max_retries}): {e}")
                    if attempt < max_retries - 1:
                        await asyncio.sleep(2 ** attempt)
                except TelegramError as e:
                    logger.error(f"[ERROR] Telegram API error: {e}")
                    return False
                except Exception as e:
                    logger.error(f"[ERROR] Unexpected error (attempt {attempt + 1}/{max_retries}): {e}")
                    if attempt < max_retries - 1:
                        await asyncio.sleep(2 ** attempt)
        
        except Exception as e:
            logger.error(f"[FORMAT_ERROR] Error formatting performance message: {e}")
        
        return False
    
    def process_new_events_sync(self):
        """Synchronous wrapper for async event processing"""
        try:
            asyncio.run(self._process_new_events_async())
        except Exception as e:
            logger.error(f"[CRITICAL] Error in event processing: {e}")
            self.consecutive_failures += 1
        finally:
            self.is_processing = False
    
    async def _process_new_events_async(self):
        """Main async event processing logic"""
        if self.is_processing:
            logger.warning("[SKIP] Previous check still running")
            return
        
        self.is_processing = True
        self._reset_cycle_tracking()
        start_time = datetime.now()
        
        try:
            logger.info("="*60)
            logger.info("Starting event check cycle")
            logger.info("="*60)
            
            # Process speeding events
            logger.info("[FETCH] Fetching speeding events...")
            speeding_events, speed_error = self.fetch_speeding_events()
            
            if speed_error:
                logger.error(f"[ERROR] Speeding fetch error: {speed_error}")
            
            if speeding_events is not None:
                new_speeding = self.filter_new_speeding_events(speeding_events)
                if new_speeding:
                    logger.info(f"[PROCESS] Processing {len(new_speeding)} speeding events")
                    for event in new_speeding:
                        self._mark_processed(event.get('id', 0))
                        self.store_speeding_event(event)
                        await self.send_speeding_event_to_telegram(event)
                        await asyncio.sleep(1)
                else:
                    logger.info("[PROCESS] No new speeding events")
            else:
                logger.warning("[FETCH] Failed to fetch speeding events")
            
            # Process performance events
            logger.info("[FETCH] Fetching performance events...")
            performance_events, perf_error = self.fetch_driver_performance_events()
            
            if perf_error:
                logger.error(f"[ERROR] Performance fetch error: {perf_error}")
            
            if performance_events is not None:
                new_performance = self.filter_new_performance_events(performance_events)
                if new_performance:
                    logger.info(f"[PROCESS] Processing {len(new_performance)} performance events")
                    for event in new_performance:
                        self._mark_processed(event.get('id', 0))
                        self.store_performance_event(event)
                        await self.send_performance_event_to_telegram(event)
                        await asyncio.sleep(1)
                else:
                    logger.info("[PROCESS] No new performance events")
            else:
                logger.warning("[FETCH] Failed to fetch performance events")
            
            # Update health
            self.last_successful_check = datetime.now()
            self.consecutive_failures = 0
            self.critical_alert_sent = False
            
            duration = (datetime.now() - start_time).total_seconds()
            logger.info("="*60)
            logger.info(f"Event check completed in {duration:.1f}s")
            logger.info("="*60)
            
        except Exception as e:
            logger.error(f"[CRITICAL] Unhandled error: {e}")
            import traceback
            logger.error(traceback.format_exc())
            self.consecutive_failures += 1
    
    async def send_daily_report(self):
        """Send daily Excel report at end of day"""
        try:
            logger.info("[DAILY] Preparing daily Excel report...")
            
            with self.data_lock:
                if not self.today_events:
                    logger.info("[DAILY] No events recorded today")
                    await self.telegram_bot.send_message(
                        chat_id=self.chat_id,
                        text="📋 Daily Report: No safety events recorded today."
                    )
                    return
                
                # Create Excel file
                excel_file = self.create_excel_file()
            
            if excel_file and os.path.exists(excel_file):
                await self.send_excel_file(excel_file, f"Daily Safety Report - {datetime.now().strftime('%m/%d/%Y')}")
                
                # Clean up file
                try:
                    os.remove(excel_file)
                except:
                    pass
                
                # Reset today's events
                with self.data_lock:
                    self.today_events.clear()
                
                logger.info("[DAILY] Daily report sent successfully")
            else:
                logger.error("[DAILY] Failed to create Excel file")
        
        except Exception as e:
            logger.error(f"[DAILY_ERROR] Error sending daily report: {e}")
    
    async def handle_getexcel_command(self):
        """Send today's Excel report when /getexcel command is used"""
        try:
            logger.info("[COMMAND] /getexcel command received")
            
            with self.data_lock:
                if not self.today_events:
                    logger.info("[COMMAND] No events for today")
                    await self.telegram_bot.send_message(
                        chat_id=self.chat_id,
                        text="📋 Today's Report: No safety events recorded so far."
                    )
                    return
                
                excel_file = self.create_excel_file()
            
            if excel_file and os.path.exists(excel_file):
                await self.send_excel_file(excel_file, f"Safety Report - {datetime.now().strftime('%m/%d/%Y')}")
                
                # Clean up file
                try:
                    os.remove(excel_file)
                except:
                    pass
                
                logger.info("[COMMAND] Excel file sent successfully")
            else:
                logger.error("[COMMAND] Failed to create Excel file")
        
        except Exception as e:
            logger.error(f"[COMMAND_ERROR] Error handling /getexcel command: {e}")
    
    async def health_check(self):
        """Perform health check and send status"""
        try:
            logger.info("[HEALTH] Running health check...")
            
            speeding_test, _ = self.fetch_speeding_events()
            performance_test, _ = self.fetch_driver_performance_events()
            apis_ok = (speeding_test is not None and performance_test is not None)
            
            telegram_ok = False
            try:
                await self.telegram_bot.get_me()
                telegram_ok = True
            except Exception as e:
                logger.error(f"[HEALTH] Telegram check failed: {e}")
            
            overall_ok = apis_ok and telegram_ok
            status = "✅ Healthy" if overall_ok else "❌ Issues Detected"
            
            last_check_str = "Never" if not self.last_successful_check else f"{(datetime.now() - self.last_successful_check).total_seconds() / 60:.1f}m ago"
            
            with self.data_lock:
                events_count = len(self.today_events)
            
            health_msg = f"""Health Report
Status: {status}
Last Check: {last_check_str}
Failures: {self.consecutive_failures}
API: {'✅ OK' if apis_ok else '❌ FAILED'}
Telegram: {'✅ OK' if telegram_ok else '❌ FAILED'}
Today's Events: {events_count}
Interval: {self.check_interval // 60}m
Version: 2.2 Excel"""
            
            await self.telegram_bot.send_message(
                chat_id=self.chat_id,
                text=health_msg,
                parse_mode='Markdown'
            )
            
            logger.info(f"[HEALTH] Report sent - Status: {status}")
            
        except Exception as e:
            logger.error(f"[HEALTH] Health check failed: {e}")
    
    def run_scheduler(self):
        """Main scheduler loop - runs in main thread"""
        # Schedule event checks
        schedule.every(self.check_interval).seconds.do(self.process_new_events_sync)
        logger.info(f"[SCHEDULER] Event check every {self.check_interval}s ({self.check_interval/60:.1f}m)")
        
        # Schedule health checks
        schedule.every().hour.do(lambda: asyncio.run(self.health_check()))
        logger.info("[SCHEDULER] Health check every hour")
        
        # Schedule daily report at end of day (11:59 PM)
        schedule.every().day.at("23:59").do(lambda: asyncio.run(self.send_daily_report()))
        logger.info("[SCHEDULER] Daily report scheduled for 23:59")
        
        # Schedule /getexcel command handler - every day at 9 AM
        schedule.every().day.at("09:00").do(lambda: asyncio.run(self.handle_getexcel_command()))
        logger.info("[SCHEDULER] /getexcel command scheduled for 09:00")
        
        # Run initial check
        logger.info("[STARTUP] Running initial event check...")
        self.process_new_events_sync()
        
        # Main loop
        logger.info("[STARTUP] SafetyBot v2.2 Excel is now monitoring")
        print("\n" + "="*60)
        print("SafetyBot v2.2 Excel - Production Ready")
        print(f"Check interval: {self.check_interval // 60} minutes")
        print(f"Monitoring: Speeding & 6 Performance Events")
        print("Daily Excel report at 23:59")
        print("Press Ctrl+C to stop")
        print("="*60 + "\n")
        
        while self.running:
            try:
                schedule.run_pending()
                time.sleep(10)
            except KeyboardInterrupt:
                logger.info("[SHUTDOWN] Bot stopped by user")
                break
            except Exception as e:
                logger.error(f"[ERROR] Scheduler loop error: {e}")
                time.sleep(30)
    
    def start(self):
        """Start the bot"""
        try:
            logger.info("[STARTUP] Testing connections...")
            asyncio.run(self._test_connections())
            
            self.run_scheduler()
            
        except Exception as e:
            logger.error(f"[CRITICAL] Fatal error during startup: {e}")
            sys.exit(1)
    
    async def _test_connections(self):
        """Test API and Telegram connections"""
        try:
            logger.info("[TEST] Testing Telegram connection...")
            bot_info = await self.telegram_bot.get_me()
            logger.info(f"[TEST] Telegram OK: @{bot_info.username}")
            
            await self.telegram_bot.send_message(
                chat_id=self.chat_id,
                text="✅ SafetyBot v2.2 Excel Started\nCheck Interval: " + str(self.check_interval // 60) + "m\nReady to monitor",
                parse_mode='Markdown'
            )
            logger.info("[TEST] Startup message sent")
            
        except Exception as e:
            logger.error(f"[TEST] Telegram connection failed: {e}")
            raise
        
        try:
            logger.info("[TEST] Testing API connections...")
            speeding, _ = self.fetch_speeding_events()
            performance, _ = self.fetch_driver_performance_events()
            
            s_count = len(speeding) if speeding else 0
            p_count = len(performance) if performance else 0
            
            logger.info(f"[TEST] API OK: {s_count} speeding, {p_count} performance events")
            
        except Exception as e:
            logger.error(f"[TEST] API connection failed: {e}")
            raise

def main():
    """Main entry point"""
    try:
        bot = SafetyBot()
        bot.start()
    except KeyboardInterrupt:
        logger.info("[MAIN] Bot stopped by user")
    except Exception as e:
        logger.error(f"[MAIN] Fatal error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()